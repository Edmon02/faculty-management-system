from flask import Flask, render_template, request, redirect, url_for, session, make_response, flash, get_flashed_messages, jsonify, g#, send_from_directory, send_file,
from flask_wtf import FlaskForm
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from wtforms import StringField, PasswordField, SubmitField, SelectField, FileField
from wtforms.validators import DataRequired
import random
import string
import hashlib
import secrets
import base64
from io import BytesIO
import os
#from werkzeug.utils import secure_filename
from openpyxl import Workbook
import mimetypes
import sqlite3
from functools import cache
from datetime import datetime, timedelta
import urllib.parse
from collections import defaultdict
from flask_socketio import SocketIO, emit
import time
import plotly.express as px
import pandas as pd

# Define login form
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

class AddSubjectForm(FlaskForm):
    first_name = StringField('first_name', validators=[DataRequired()])
    group = SelectField('group', choices=[('1', '1'), ('2', '2'), ('3', '3')], validators=[DataRequired()])
    file = FileField('file', validators=[DataRequired()])
    submit = SubmitField('Submit')


def random_numbers(length):
    return ''.join(str(random.randint(0, 9)) for _ in range(length))

def random_password(length):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

import sqlite3

def insert_student_data(data):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    query = """
    INSERT INTO Student (group_name, first_name, last_name, patronymic, birthday_date, phone, image)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """
    query2 = """
    INSERT INTO Users (username, password, _id)
    VALUES (?, ?, ?)
    """

    # Assuming 'image' should be a placeholder for the student's photo
    # Modify this part based on your table structure
    data['image'] = b''

    cursor.execute(query, (
        data['group_name'],
        data['first_name'],
        data['last_name'],
        data['patronymic'],
        data['birthday_date'].strftime('%Y-%m-%d'),
        data['phone'],
        data['image']
    ))

    user_id = cursor.lastrowid
    cursor.execute(query2, (
        random_numbers(8),
        hashlib.sha256(random_password(10).encode()).hexdigest(),
        user_id
    ))

    conn.commit()
    cursor.close()
    conn.close()


UPLOAD_FOLDER = 'c:\\Users\\edmon\\Downloads\\Telegram Desktop\\Polytech_Univ\\'

app = Flask(__name__)

secret_key = secrets.token_hex(16)
app.secret_key = secret_key

cur_dir = os.path.dirname(__file__)
# Define the SQLite database path
DATABASE =  os.path.join(cur_dir, 'pulpit.sqlite')

# Set secure configurations
# app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['STATIC_FOLDER'] = '/static/images'
app.config['SECRET_KEY'] = secret_key
app.config['DATABASE'] = DATABASE



# Set up rate limiting
limiter = Limiter(
    key_func=get_remote_address,     
    storage_uri="memory://",
    )
# limiter = Limiter(key_func=get_remote_address, storage_uri="redis://localhost:5000")
limiter.init_app(app)

# Set session configurations
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = '/tmp/flask_session'

# Set the lifetime of a permanent session
app.permanent_session_lifetime = 60 * 60 * 12  # half day
socketio = SocketIO(app)

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
    return g.db

# Function to check and invalidate previous session
def check_and_invalidate_session(user_id, current_ip):
    if 'user_id' in session:
        if session['ID'] == user_id and session.get('user_ip') != current_ip:
            # The user is logging in from a new IP, invalidate previous session
            session.clear()
            return True
        else :
            session['ID'] = user_id
            session['user_ip'] = current_ip
    return False

# Define a list of restricted routes for students
# Register the custom 404 error handler
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

restricted_routes_for_students = ['/students/add-student']

# Define a before_request function to check and redirect if needed
@app.before_request
def check_user_access():
    # Get the requested path
    requested_path = request.path
    # Check if the user is a student and the requested path is restricted
    is_student = False  # Replace this with your logic to determine if the user is a student
    if is_student and requested_path in restricted_routes_for_students:
        # Redirect the student to a different page, e.g., a 404 page
        return render_template('404.html'), 404

@app.route('/')
def news():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM News ORDER BY date DESC')
    column_names = [description[0] for description in cursor.description]
    news = [dict(zip(column_names, row)) for row in cursor.fetchall()]
    # Convert image data to base64
    for item in news:
        if 'cover' in item:
            item['cover'] = base64.b64encode(item['cover']).decode('utf-8')

    conn.commit()
    conn.close()
    return render_template('index.html', news=news)


@app.route('/', methods=['GET', 'POST'])
def index():
    form = LoginForm()
    if bool(form.username.data and form.password.data) and 'logged_in' not in session:
        data = request.get_json()

        # Extract username and password from the request
        username = data.get('username')
        password = data.get('password')
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        query = "SELECT * FROM Users WHERE username = ? AND password = ?"

        cursor.execute(query, (username, password))

        # Fetch the first row
        result = cursor.fetchone()
        
        if result is not None:
            column_names = [description[0] for description in cursor.description]

            # Fetch the results
            user_data = dict(zip(column_names, result))
            query = "SELECT * FROM Student WHERE id = ?" if user_data['type'] == 'student' else "SELECT * FROM Lecturer WHERE id = ?"

            cursor.execute(query, (user_data['_id'], ))
            # Get the column names
            column_names = [description[0] for description in cursor.description]

            # Fetch the results
            data = dict(zip(column_names, cursor.fetchone()))
            # print(data)

            if user_data['type'] != 'student':
                query = """
                        SELECT
                            group_name
                        FROM LecturerGroup lg
                        WHERE lg.lecturer_id = ?;
                        """
                cursor.execute(query, (data['id'],))
                # Fetch the results
                group_ids = [result[0] for result in cursor.fetchall()]
            else:
                group_ids = str(data['group_name'])

        cursor.close()
        conn.close()
        if result is not None and data:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            # if there is activity for the current month
            user_id = data['id']
            current_date = datetime.now()
            cursor.execute("SELECT date FROM ActivityLog WHERE  strftime('%Y-%m', date) = ?", (current_date.strftime('%Y-%m'), ))
            activity_log_id = cursor.fetchone()

            if activity_log_id:
                # Update activity_count in the User table
                cursor.execute("UPDATE ActivityLog SET activity_count = activity_count + 1 WHERE strftime('%Y-%m', date) = ?", (current_date.strftime('%Y-%m'),))
            else:
                # Insert a new record in the ActivityLog table
                cursor.execute("INSERT INTO ActivityLog (date) VALUES (?)", (current_date.strftime('%Y-%m-%d'), ))

            conn.commit()
            conn.close()
            # student_id = data[0]
            # hashed_password = data['password']
            current_ip = request.remote_addr
            # # Check and invalidate previous session if necessary
            # if check_and_invalidate_session(data['id'], current_ip):
            #     return jsonify({'message': 'Previous session invalidated. Please log in again.'})
            session['logged_in'] = True
            session['type'] = user_data['type']
            session['first_name'] = data['first_name']
            session['last_name'] = data['last_name']
            session['is_Admin'] = data['is_Admin']
            session['is_Lecturer'] = data['is_Lecturer']
            session['ID'] = user_id
            session['user_ip'] = current_ip

            session['group_ids'] = group_ids if isinstance(group_ids, list) else (group_ids,)
            return jsonify({'message': 'Login successful'})
    elif 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html', form=form)

def get_user_activity_data():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # Example SQL query, replace with your actual query
    cursor.execute("SELECT date, activity_count FROM ActivityLog")
    results = cursor.fetchall()

    # Convert results to a list of dictionaries
    user_data = [{'date': date, 'activity_count': activity_count} for date, activity_count in results]

    cursor.close()
    conn.close()

    return user_data

@app.route('/logout')
def logout():
    # Clear the session data
    session.clear()
    # Redirect to the login page
    return redirect(url_for('index'))

@cache
@app.route('/dashboard')
def dashboard():
    if 'logged_in' in session:
        ID = session.get('ID')
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # # Retrieve student data from SQLite
        # query = "SELECT * FROM Student WHERE id = ? "
        # cursor.execute(query, (ID, ))
        # # Get the column names
        # column_names = [description[0] for description in cursor.description]

        # # Fetch the results
        # data = [dict(zip(column_names, cursor.fetchone()))]

        group_ids = tuple(session['group_ids'])
        query = """
        SELECT *
        FROM Student
        WHERE group_name IN ({})
        """.format(','.join('?' for _ in group_ids))

        cursor.execute(query, group_ids)
        column_names = [description[0] for description in cursor.description]

        rows = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        # rows = cursor.fetchall()
        group_data = {}
        for row in rows:
            group_name = row['group_name']

            if group_name not in group_data:
                group_data[group_name] = {
                    'group_name': group_name,
                    'student_datas': [row]
                }
            else:
                group_data[group_name]['student_datas'].append(row)

        data = {}
        data['group_datas'] = group_data

        if session['type'] == 'student':

            # conn2 = sqlite3.connect(DATABASE)
            # cursor2 = conn2.cursor()

            query = """
            SELECT s.subject_id, s.subject_name
            FROM Subject s
            WHERE s.subject_id IN (
                SELECT slg.subject_id
                FROM SubjectLecturerGroup slg
                WHERE slg.lecturer_id IN (
                    SELECT lecturer_id
                    FROM Lecturer
                    WHERE lecturer_id IN (
                        SELECT lecturer_id
                        FROM SubjectLecturerGroup
                        WHERE group_name in ({})
                    )
                )
            );
            """.format(','.join('?' for _ in group_ids))

            cursor.execute(query, group_ids)
            column_names2 = [description[0] for description in cursor.description]
            subject_data = [dict(zip(column_names2, row)) for row in cursor.fetchall()]
            file_datas = []
            for subject in subject_data:
                subject_id = subject['subject_id']
                query = """
                    SELECT * 
                    FROM Files f
                    WHERE f.id IN (
                        SELECT sf.file_id
                        FROM SubjectFile sf
                        WHERE sf.subject_id = ?
                        )"""
                cursor.execute(query, (subject_id,))
                column_names = [description[0] for description in cursor.description]
                files = [dict(zip(column_names, row)) for row in cursor.fetchall()]
                file_data = [{
                    'file_name': file['file_name'],
                    'file_type': os.path.splitext(file['file_name'])[1],
                    'file_id': file['id']
                } for file in files]
                subject_data = {
                    'subject_name': subject['subject_name'],
                    'files': file_data
                }
                file_datas.append(subject_data)
            data['files'] = file_datas
            cursor.close()


        conn.close()
        # Replace this with your actual data retrieval logic
        user_data = get_user_activity_data()

        # Create a Plotly figure
        fig = px.line(user_data, x='date', y='activity_count', labels={'activity_count': 'User Activity'})

        # You can customize the layout if needed
        fig.update_layout(title='User Activity Over Time', xaxis_title='Date', yaxis_title='Activity Count')

        # Convert the Plotly figure to JSON for rendering in the template
        graph_json = fig.to_json()


        return render_template('dashboard.html', data=data, graph_json=graph_json)
    else:
        return redirect(url_for('index'))

@app.route('/students', methods=['GET', 'POST'])
def studentsList():
    if 'logged_in' in session:
        group_ids = session.get('group_ids')

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Retrieve student data from SQLite based on group_id
        query = "SELECT * FROM Student WHERE group_name in ({})".format(','.join('?' for _ in group_ids))
        cursor.execute(query, group_ids)
        # Get the column names
        column_names = [description[0] for description in cursor.description]

        # Fetch the results
        data_student = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        conn.close()

        if request.method == 'GET':
            search_params = {}

            # Extract query parameters and update search_params dictionary
            search_params['id'] = request.args.get('id', default='', type=str)
            search_params['first_name'] = request.args.get('first_name', default='', type=str)
            search_params['phone'] = request.args.get('phone', default='', type=str)

            # Filter Data_student based on search_params
            if search_params['id']:
                data_student = [data for data in data_student if search_params['id'] in data['_id']]
            if search_params['first_name']:
                data_student = [data for data in data_student if search_params['first_name'].lower() == data['first_name'].lower()]
            if search_params['phone']:
                data_student = [data for data in data_student if search_params['phone'] in data['phone']]


        if request.method == 'POST' and request.form['action'] == 'Sort':
            data_student = sorted(data_student, key=lambda x: x['rating'], reverse=True)

        return render_template('students.html', data=data_student)
    else:
        return redirect(url_for('index'))

@app.route('/students/add-student', methods=['GET', 'POST'])
def addStudent():
    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        dob = request.form['dob']
        group_name = request.form['group']
        patronymic = request.form['patronymic']
        email = request.form['email']
        phone = request.form['phone']
        photo = request.files['photo']
        # Handle file upload
        excel_file = request.files['excel_file']

        if excel_file:
            # Read the Excel file using pandas
            df = pd.read_excel(excel_file)
            # Iterate through rows and insert into the SQLite table
            for _, row in df.iterrows():
                print(row['Birthday'].to_pydatetime())
                # Extract data from each row
                student_data = {
                    'first_name': row['Name'],
                    'last_name': row['Surname'],
                    'patronymic': row['Patronymic'],
                    'phone': row['Phone'],
                    'address': '',
                    'birthday_date': row['Birthday'].to_pydatetime(),
                    'group_name': row['Group'],
                    # Add other columns as needed
                }

                # Insert the data into the SQLite table
                insert_student_data(student_data)

            return redirect(url_for('studentsList'))


        print(request.url)
        if photo:
            picture_data = photo.read()
        else:
            # Open the image photo and read its contents
            with open("images.jpg", "rb") as f:
                image_data = f.read()

            # Encode the image data as base64
            picture_data = base64.b64encode(image_data)
            # If no photo was uploaded, use a default photo
            # db.students.insert_one({'photo_path': 'default.png'})
        print(dob)
        student_data = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'patronymic': patronymic,
                    'phone': phone,
                    'address': '',
                    'birthday_date': datetime.strptime(dob, '%Y-%m-%d'),
                    'group_name': group_name,
                    # Add other columns as needed
                }

        insert_student_data(student_data)

        return redirect(url_for('dashboard'))
    else:
        return render_template('add-student.html')

@app.route('/students/edit-student/<int:id>', methods=['GET', 'POST'])
def editStudent(id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    query = "SELECT * FROM Student WHERE id = ?"
    cursor.execute(query, (id,))
    # Get the column names
    column_names = [description[0] for description in cursor.description]

    # Fetch the results
    data_cursor = dict(zip(column_names, cursor.fetchone()))

    if request.method == 'POST':
        search_params = dict(data_cursor)

        # Extract form data and update search_params dictionary
        search_params['last_name'] = request.form['last_name']
        search_params['first_name'] = request.form['first_name']
        search_params['patronymic'] = request.form['patronymic']
        search_params['birthday_date'] = request.form['dob']
        search_params['group_name'] = request.form.get('group_name', default=data_cursor['group_name'], type=int)
        search_params['phone'] = request.form['phone']

        # Update the student record in the database
        old_data = data_cursor
        if old_data != search_params:
            update_query = "UPDATE Student SET last_name = ?, first_name = ?, patronymic = ?, birthday_date = ?, group_name = ?, phone = ? WHERE id = ?"
            cursor.execute(update_query, (
                search_params['last_name'],
                search_params['first_name'],
                search_params['patronymic'],
                search_params['birthday_date'],
                search_params['group_name'],
                search_params['phone'],
                id
            ))
            conn.commit()
            archive_query = "INSERT INTO Archive_student (id, last_name, first_name, patronymic, birthday_date, group_name, phone) VALUES (?, ?, ?, ?, ?, ?, ?)"
            cursor.execute(archive_query, (
                old_data['id'],
                old_data['last_name'],
                old_data['first_name'],
                old_data['patronymic'],
                old_data['birthday_date'],
                old_data['group_name'],
                old_data['phone']
            ))
            conn.commit()

        # Redirect the user to the student list page
        return redirect(url_for('studentsList'))

    conn.close()

    return render_template('edit-student.html', data=data_cursor)


@app.route('/file')
def show_file():
    # Get the name of the file to download from the query string parameter
    filename = request.args.get('filename')

    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Retrieve the file data from SQLite
        query = "SELECT file FROM Files WHERE file_name = ?"
        cursor.execute(query, (filename,))
        file_data = cursor.fetchone()

        conn.close()

        if file_data:
            file_data = file_data[0]  # Encode as bytes
            contents = BytesIO(file_data)

            # Set the Content-Type header based on file extension
            content_type, _ = mimetypes.guess_type(filename)
            if not content_type:
                content_type = 'application/octet-stream'

            response = make_response(contents.getvalue())
            response.headers.set('Content-Type', content_type)

            # Set the Content-Disposition header
            filename_header = 'filename="%s"' % urllib.parse.quote(filename.encode('utf-8'))
            if content_type != 'application/pdf':
                response.headers.set('Content-Disposition', 'attachment', filename=filename_header)
            else:
                response.headers.set('Content-Disposition', 'inline', filename=filename_header)

            return response
        else:
            return "File not found."
    except Exception as e:
        return f"An error occurred: {str(e)}"

@app.route('/teachers', methods=['GET', 'POST'])
def teachers():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # query = 'SELECT chair_code FROM groups WHERE _id = ?'
    # group_id = session.get('group_id')
    # cursor.execute(query, (group_id,))
    # group_data = cursor.fetchone()
    group_ids = session['group_ids']

    query = """
        SELECT * FROM Lecturer l
        WHERE l.id IN (
            SELECT lecturer_id
            FROM LecturerGroup
            WHERE group_name in ({})
        )
        """.format(','.join('?' for _ in group_ids))
    cursor.execute(query, group_ids)

    # Get the column names
    column_names = [description[0] for description in cursor.description]

    # Fetch the results
    lecturer_data = [dict(zip(column_names, row)) for row in cursor.fetchall()]
    if request.method == 'GET':
        search_params = {}

        # Extract query parameters and update search_params dictionary
        search_params['id'] = request.args.get('id', default='', type=str)
        search_params['surname'] = request.args.get('surname', default='', type=str)
        search_params['phone'] = request.args.get('phone', default='', type=str)

        # Filter lecturer_data based on search_params
        if search_params['id']:
            lecturer_data = [data for data in lecturer_data if search_params['id'] in data['_id']]
        if search_params['surname']:
            lecturer_data = [data for data in lecturer_data if search_params['surname'].lower() == data['surname'].lower()]
        if search_params['phone']:
            lecturer_data = [data for data in lecturer_data if search_params['phone'] in data['phone']]

    if request.method == 'POST' and request.form['action'] == 'Download':
        # Create a new Excel workbook and sheet
        wb = Workbook()
        sheet = wb.active

        # Add column headers to the sheet
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Name'
        sheet['C1'] = 'Surname'
        sheet['D1'] = 'Email'
        sheet['E1'] = 'Degree'

        # Add data to the sheet
        for i, data in enumerate(lecturer_data, start=2):
            sheet.cell(row=i, column=1, value=data['id'])
            sheet.cell(row=i, column=2, value=data['first_name'])
            sheet.cell(row=i, column=3, value=data['last_name'])
            sheet.cell(row=i, column=4, value=data['email'])
            sheet.cell(row=i, column=5, value=data['academic_degree'])

        # Create a bytes buffer to save the workbook to
        buffer = BytesIO()

        # Save the workbook to the buffer
        wb.save(buffer)

        # Set the buffer's seek pointer to the beginning
        buffer.seek(0)

        # Create a response object with the workbook as the content and headers to trigger a file download
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=teachers.xlsx'

        conn.close()
        return response

    conn.close()
    return render_template('teachers.html', data=lecturer_data)

def add_file(subject_id, file_name):
    check_subj = "SELECT COUNT(*) FROM SubjectFile WHERE subject_id AND "


    # Check if the record already exists
    check_query = "SELECT COUNT(*) FROM Files WHERE subject_id = ? AND file_name = ?"
    cursor.execute(check_query, (subject_id, file_name))
    record_exists = cursor.fetchone()
    if not record_exists[-1]:
        add_file_query = "INSERT OR IGNORE INTO Files (subject_id, file, file_name) VALUES (?, ?, ?)"
        cursor.execute(add_file_query, (subject_id, file, file_name))
        record_exists = cursor.lastrowid
    
    add_subj_file = "INSERT OR"

@app.route('/subjects', methods=['GET', 'POST'])
def subjects():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    # group_ids = session['group_ids']
    group_ids = ['Group A', 'Group C']

    query2 = """
    SELECT s.subject_id, s.subject_name
    FROM Subject s
    WHERE s.subject_id IN (
         SELECT slg.subject_id
         FROM SubjectLecturerGroup slg
         WHERE slg.lecturer_id IN (
              SELECT lecturer_id
              FROM Lecturer
              WHERE lecturer_id IN (
                   SELECT lecturer_id
                   FROM SubjectLecturerGroup
                   WHERE group_name in ({})
              )
         )
    );
    """.format(','.join('?' for _ in group_ids))
    
    cursor.execute(query2, group_ids)
    column_names2 = [description[0] for description in cursor.description]
    subject_data = [dict(zip(column_names2, row)) for row in cursor.fetchall()]
    
    data = []
    for subject in subject_data:
        subject_id = subject['subject_id']
        query = """
            SELECT * 
            FROM Files f
            WHERE f.id IN (
                SELECT sf.file_id
                FROM SubjectFile sf
                WHERE sf.subject_id = ?
                )"""
        cursor.execute(query, (subject_id,))
        column_names = [description[0] for description in cursor.description]
        files = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        file_data = [{
            'file_name': file['file_name'],
            'file_type': os.path.splitext(file['file_name'])[1],
            'subject_id': subject_id,
            'file_id': file['id']
        } for file in files]
        subject_data = {
            'subject_name': subject['subject_name'],
            'files': file_data
        }
        data.append(subject_data)

    cursor.close()
    conn.close()

    return render_template('subjects.html', file_data = data)


@app.route('/add-subject', methods=['GET', 'POST'])
def addSubjects():
    if request.method == 'POST':
        end_time = int(request.form['end_time'])
        group_name = request.form['group_name']
        subject_name = request.form['subject_name']

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        expiry_time = datetime.now() + timedelta(seconds=end_time)
        datas = (group_name, messenge, subject_name)
        query = "INSERT INTO Subject (group_name, messenge, subject_name) VALUES (?, ?, ?)"
        cursor.execute(query, datas)
        conn.commit()
        conn.close()

    else :
        return render_template('add-subject.html')
    # form = AddSubjectForm()
    # if form.validate_on_submit():
    #     # Get form data
    #     file = request.files['file']
    #     file_data = file.read()

    #     # create a binary object from the file data
    #     binary_data = bson.binary.Binary(file_data)

    #     first_name = request.form['first_name']
    #     group = request.form['group']

    #     # Store data in MongoDB
    #     document =  {
    #         'first_name': first_name,
    #         'group_id': int(group),
    #         'file_name': file.filename,
    #         'file':binary_data
    #     }
    #     # Save data to MongoDB
    #     files.insert_one(document)

    #     flash('Subject added successfully!', 'success')
    #     time.sleep(5) # Delay for 2 seconds
    #     # return redirect(url_for('dashboard'))

    # # Check if form has been submitted and all fields are not filled
    # if request.method == 'POST' and not form.validate():
    #     flash('Please fill out all fields.', 'danger')

    # messages = get_flashed_messages()
    # if messages:
    #     return render_template('add-subject.html', form=form, messages=messages)

    #     # # Check file size
    #     # file = request.files['file']
    #     # file_size = len(file.read())
    #     # if file_size > 10485760:  # 10 MB in bytes
    #     #     return render_template('add-subject.html', error='File size exceeds the limit')
    #     # filename = file.filename
    #     # print(filename, file_size)
    return render_template('add-subject.html')

@app.route('/delte-subject/<int:id1>/<int:id2>', methods=['GET', 'POST'])
def deleteSubject(id1, id2):
    if request.method == 'GET':
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        query = "DELETE FROM SubjectFile WHERE subject_id = ? and file_id = ?"
        cursor.execute(query, (id1, id2))
        conn.commit()
        conn.close()
    return redirect(url_for('subjects'))

@app.route("/exercises", methods=['POST', 'GET'])
def exercie():
    group_ids = session['group_ids']
    # group_ids = (1, )
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    query = """
        SELECT
            *
        FROM Exercises
        WHERE id in (
            SELECT exercise_id
            FROM ExercisesGroup
            WHERE group_name in ({})
        )
    """.format(','.join('?' for _ in group_ids))

    cursor.execute(query, group_ids)

    column_names = [description[0] for description in cursor.description]
    # Fetch the results
    exercise_data = [dict(zip(column_names, row)) for row in cursor.fetchall()]
    checkTypes = []
    if session['type'] == 'student':
        query = '''SELECT exercise_id
                    FROM WaitRoom
                    Where user_id = ?    
                '''
        cursor.execute(query, (session['ID'],))
        checkTypes = [row[0] for row in cursor.fetchall()]
    print(checkTypes)

    for exercise in exercise_data:
        exercise['checkType'] = 1 if exercise['id'] in checkTypes else 0

    exercise_data = [{
        'id': data['id'],
        'group_name': data['group_name'],
        'expiry_time': data['expiry_time'],
        'subject_name': data['subject_name'],
        'messenge': data['messenge'],
        'checkType': data['checkType'],
        'file_name': data['file_name'],
        'file_type': os.path.splitext(data['file_name'])[1]
    } for data in exercise_data]
    
    return render_template('exercises.html', data=exercise_data)

@app.route('/add-exercise', methods=['GET', 'POST'])
def addExercise():
    if request.method == 'POST':
        end_time = datetime.strptime(request.form['end_time'], '%Y-%m-%d')
        group_name = request.form['group_name']
        messenge = request.form['messenge']
        subject_name = request.form['subject_name']
        file = request.files['file'].read()
        file_name = request.files['file'].filename
        group_id = 1

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        subject_id_query = "SELECT subject_id FROM Subject WHERE subject_name = ?"
        cursor.execute(subject_id_query, (subject_name, ))
        subject_id = cursor.fetchone()[0]

        add_file(subject_i, file_name)

        # print(end_time)
        delta = datetime.now() - end_time
        expiry_time = (datetime.now() + delta).strftime("%Y-%m-%d")
        datas = (expiry_time, group_name, messenge, subject_name, file_name, group_id)
        query = "INSERT INTO Exercises (expiry_time, group_name, messenge, subject_name, file_name, group_id) VALUES (?, ?, ?, ?, ?, ?)"
        # query1 = "INSERT INTO ExerciseStudent (group_id) VALUES (?)"
        cursor.execute(query, datas)
        # cursor.execute(query1, datas)
        conn.commit()
        conn.close()

    return render_template('add-exercise.html')

@app.route('/edit-exercise/<int:id>', methods=['GET', 'POST'])
def editExercise(id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    query = "SELECT * FROM Exercises WHERE id = ?"
    cursor.execute(query, (id,))
    # Get the column names
    column_names = [description[0] for description in cursor.description]

    # Fetch the results
    data_cursor = dict(zip(column_names, cursor.fetchone()))
    expiry_time = datetime.strptime(data_cursor['expiry_time'], "%Y-%m-%d")

    data_cursor['expiry_time'] = "{:%Y-%m-%d}".format(expiry_time)
    if request.method == 'POST':
        search_params = dict(data_cursor)
        filename = request.files['file_name'].filename

        # Extract form data and update search_params dictionary
        search_params['group_name'] = request.form['group_name']
        search_params['expiry_time'] = request.form['expiry_time']
        search_params['subject_name'] = request.form['subject_name']
        search_params['messenge'] = request.form.get('messenge')
        search_params['file_name'] = data_cursor['file_name'] if filename == '' else filename


        # Update the exercise record in the database
        old_data = data_cursor
        if old_data != search_params:
            update_query = "UPDATE Exercises SET group_name = ?, expiry_time = ?, subject_name = ?, messenge = ?, file_name = ? WHERE id = ?"
            cursor.execute(update_query, (
                search_params['group_name'],
                search_params['expiry_time'],
                search_params['subject_name'],
                search_params['messenge'],
                search_params['file_name'],
                id
            ))
            conn.commit()

        # Redirect the user to the exercise list page
        return redirect(url_for('exercie'))

    conn.close()

    return render_template('edit-exercise.html', data=data_cursor)

@app.route('/delete-exercise/<int:id>', methods=['GET', 'POST'])
def deleteExercise(id):
    if request.method == 'GET':
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        query = "DELETE FROM Exercises WHERE id in (?)"
        cursor.execute(query, (id, ))
        conn.commit()
        conn.close()
    return redirect(url_for('exercie'))

# Define a function to insert data into the WaitRoom table
def insert_into_waitroom(user_info, selected_ids, group_id):
    try:
        # Connect to the SQLite database (replace with your database configuration)
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Insert selected chat items into the WaitRoom table
        for chat_item_id in selected_ids:
            cursor.execute("INSERT INTO WaitRoom (user_id, group_id, exercise_id) VALUES (?, ?, ?)", (user_info, group_id, chat_item_id))
            cursor.execute("UPDATE Exercises SET checkkType = ? WHERE id = ?", (1, chat_item_id))


        # Commit the changes to the database
        conn.commit()

        # Close the database connection
        conn.close()

        return True
    except Exception as e:
        print(f"Error inserting data into WaitRoom: {str(e)}")
        return False

@app.route('/waitroom', methods=['GET', 'POST'])
def waitroom():
    if request.method == 'POST':
        # try:
            # Get the data sent in the request
            data = request.json
            # Extract user information and selected chat items
            # user_id = session['ID']
            user_id = 1
            selected_ids = data.get('selectedIds', [])
            # group_id = session['group_ids']
            group_id = 1
            # Insert data into the WaitRoom table
            if insert_into_waitroom(user_id, selected_ids, group_id):
                return jsonify({'message': 'Data sent to WaitRoom successfully'})
            else:
                return jsonify({'error': 'Failed to insert data into WaitRoom'})
        # except Exception as e:
        #     return jsonify({'error': str(e)})
    return redirect(url_for('exercie'))

@app.route('/waitingroom', methods=['GET', 'POST'])
def waitingRoom():
    # Connect to the SQLite db
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    group_ids = [1, 2, 3]

    # Retrieve data for each exercise and their associated group and student details
    query = """
        SELECT e.id as exercise_id, e.subject_name, e.group_name, wr.student_id, s.*
        FROM Exercises e
        JOIN WaitRoom wr ON e.id = wr.exercise_id
        JOIN Student s ON wr.student_id = s.id;
    """

    cursor.execute(query)
    column_names = [description[0] for description in cursor.description]
    data = [dict(zip(column_names, row)) for row in cursor.fetchall()]

    # Create a nested dictionary to store the data structure
    exercise_group_data = defaultdict(lambda: {'exercise_id': None, 'subject_name': None, 'group_ids': []})

    # Organize the data into the desired structure
    for row in data:
        exercise_id = row['exercise_id']
        subject_name = row['subject_name']
        group_name = row['group_name']
        student_id = row['student_id']
        
        # Initialize the exercise entry if it doesn't exist
        if exercise_id not in exercise_group_data:
            exercise_group_data[exercise_id]['exercise_id'] = exercise_id
            exercise_group_data[exercise_id]['subject_name'] = subject_name
        
        # Find the group entry or create a new one
        group_entry = None
        for group in exercise_group_data[exercise_id]['group_ids']:
            if group['group_name'] == group_name:
                group_entry = group
                break
        
        if group_entry is None:
            group_entry = {'group_name': group_name, 'students_data': []}
            exercise_group_data[exercise_id]['group_ids'].append(group_entry)
        
        # Add student data to the group
        student_data = {key: row[key] for key in row.keys() if key not in ('exercise_id', 'subject_name', 'group_name', 'student_id')}
        group_entry['students_data'].append({'student_id': student_id, 'student_data': student_data})
    
    cursor.close()
    conn.close()
    return render_template('waitingroom.html', datas=exercise_group_data)

@app.route('/outwaitingroom/<int:id1>/<int:id2>', methods=['GET', 'POST'])
def outWaitingRoom(id1, id2):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    print(id1, id2)
    cursor.execute("DELETE FROM WaitRoom WHERE exercise_id = ? AND student_id = ?", (id1, id2))
    conn.commit()
    conn.close()
    return redirect(url_for('waitingRoom'))

@app.route('/done/<int:exercise_id>/<int:student_id>/<int:selected_value>', methods=['GET', 'POST'])
def done(exercise_id, student_id, selected_value):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM WaitRoom WHERE exercise_id = ? AND student_id = ?", (exercise_id, student_id))
    cursor.execute("DELETE FROM ExercisesGroup WHERE exercise_id = ? AND group_name = ?", (exercise_id, 'Group A'))
    cursor.execute("UPDATE Student SET rating = rating + ? WHERE id = ?", (selected_value, student_id))
    conn.commit()
    conn.close()
    return redirect(url_for('waitingRoom'))

@app.route('/message')
def message():
    receiver_id = 1  # Replace with the actual receiver's ID
    sender_id = 2
    # Retrieve chat messages from the database for the given receiver
    db = get_db()
    cursor = db.cursor()
    query = """
        SELECT sender_id, text
        FROM Message
        WHERE receiver_id = ? OR sender_id = ?
        ORDER BY timestamp -- Replace 'timestamp_column' with your actual timestamp column name
    """
    cursor.execute(query, (receiver_id, receiver_id))
    messages = cursor.fetchall()

    return render_template('message_archive.html', current_user={'id': receiver_id}, messages=messages)


@socketio.on('message')
def handle_message(data):
    sender_id = data['sender_id']
    # receiver_id = data['receiver_id']
    # sender_id = 1
    receiver_id = 2
    message_text = data['message']
    print(sender_id, receiver_id)
    # Get the current timestamp
    timestamp = datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ')

    db = get_db()
    cursor = db.cursor()

    # Check if the sender is an Admin
    cursor.execute("SELECT is_Admin FROM Users WHERE id = ?", (sender_id,))
    sender_is_admin = cursor.fetchone()[0] == 1

    if sender_is_admin:
        # Admin can message everyone
        cursor.execute("INSERT INTO Message (text, sender_id, receiver_id, timestamp, is_group_message) VALUES (?, ?, ?, ?, ?)",
                       (message_text, sender_id, receiver_id, timestamp, 0))
    else:
        # Check if the sender is a lecturer or student
        cursor.execute("SELECT type FROM Users WHERE id = ?", (sender_id,))
        sender_role = cursor.fetchone()[0]

        if sender_role == 'lecturer':
            # Lecturers can send group messages to groups they teach
            cursor.execute("SELECT group_name FROM LecturerGroup WHERE id = ?", (sender_id,))
            lecturer_group = cursor.fetchone()[0]

            if receiver_id is None:
                # Group message
                cursor.execute("INSERT INTO Message (text, sender_id, receiver_id, timestamp, is_group_message) VALUES (?, ?, ?, ?, ?)",
                               (message_text, sender_id, None, timestamp, 1))
            else:
                # Individual message
                cursor.execute("INSERT INTO Message (text, sender_id, receiver_id, timestamp, is_group_message) VALUES (?, ?, ?, ?, ?)",
                               (message_text, sender_id, receiver_id, timestamp, 0))
        elif sender_role == 'student':
            # Students can only send messages to their group
            cursor.execute("SELECT group_name FROM Student WHERE id = ?", (sender_id,))
            student_group = cursor.fetchone()[0]

            if receiver_id is None:
                # Group message
                cursor.execute("INSERT INTO Message (text, sender_id, receiver_id, timestamp, is_group_message) VALUES (?, ?, ?, ?, ?)",
                               (message_text, sender_id, None, timestamp, 1))
            else:
                # Individual message
                cursor.execute("INSERT INTO Message (text, sender_id, receiver_id, timestamp, is_group_message) VALUES (?, ?, ?, ?, ?)",
                               (message_text, sender_id, receiver_id, timestamp, 0))

    db.commit()
    emit('message', {'message': message_text, 'sender_id': sender_id, 'timestamp': timestamp}, broadcast=True)

@app.route('/add-news', methods=['GET', 'POST'])
def addNews():
    if request.method == 'POST':
        category = request.form['category']
        title = request.form['title']
        desc = request.form['messenge']
        image = 'static/react/media/news3.2c34506ddfc7636dfa4c.jpg'
        with open(image, "rb") as f:
                image_data = f.read()

        current_datetime = datetime.now()
        formatted_date = current_datetime.strftime('%Y-%m-%d')
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        query = "INSERT INTO News (category, title, date, cover, desc) VALUES (?, ?, ?, ?, ?)"
        data = (category, title, formatted_date, sqlite3.Binary(image_data), desc)
        print(data)
        cursor.execute(query, data)

        conn.commit()
        conn.close()

    return render_template('news.html')


# # Background task to check and delete expired rows
# def delete_expired_rows():
#     # with app.app_context():
#     # Connect to the SQLite database
#     conn = sqlite3.connect(DATABASE)
#     cursor = conn.cursor()

#     current_time = datetime.now()
#     query = "DELETE FROM Exercises WHERE expiry_time <= ?"
#     cursor.execute(query, (current_time,))
#     conn.commit()
#     print("Expired values deleted.")
#     conn.commit()
#     conn.close()

# Lecturer:
# fYRKVPTdzm
# 71319352

# Admin;
# fYRKVPTdzT
# 03611557

# Student:
# ElwAiWgAZg
# 03611558


if __name__ == '__main__':
    socketio.run(app, debug=True)