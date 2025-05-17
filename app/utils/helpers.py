import base64
import mimetypes
import uuid
from datetime import datetime
from functools import wraps
from io import BytesIO

import pandas as pd
from flask import flash, make_response, redirect, session, url_for
from openpyxl import Workbook


def get_mac_address():
    """Get the MAC address of the current machine."""
    mac_address = ":".join(["{:02x}".format((uuid.getnode() >> elements) & 0xFF) for elements in range(0, 2 * 6, 2)][::-1])
    return mac_address


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:  # Assuming 'user_id' is set in session when user logs in
            flash("Please log in to access this page.", "warning")
            return redirect(url_for("auth.login"))  # Adjust 'auth.login' to your login route
        return f(*args, **kwargs)

    return decorated_function


def format_date(date_str):
    """Format a date string to a standardized format."""
    if isinstance(date_str, str):
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj.strftime("%Y-%m-%d")
        except ValueError:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S.%f")
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                return date_str
    return date_str


def encode_image_to_base64(image_data):
    """Encode binary image data to base64 string."""
    if isinstance(image_data, bytes):
        return base64.b64encode(image_data).decode("utf-8")
    return ""


def export_to_excel(data, columns):
    """
    Export data to an Excel file.

    Args:
        data: List of dictionaries with data to export
        columns: Dictionary mapping column names to display names

    Returns:
        BytesIO object with Excel file content
    """
    wb = Workbook()
    sheet = wb.active

    # Add column headers
    for col_idx, (_, header) in enumerate(columns.items(), start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Add data
    for row_idx, item in enumerate(data, start=2):
        for col_idx, key in enumerate(columns.keys(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=item.get(key, ""))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def create_file_response(buffer, filename, content_type=None):
    """Create a file download response."""
    if not content_type:
        content_type, _ = mimetypes.guess_type(filename)
        if not content_type:
            content_type = "application/octet-stream"

    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = content_type
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"

    return response


def prepare_file_response(file_data, filename):
    """Prepare a file response based on file type."""
    content_type, _ = mimetypes.guess_type(filename)
    if not content_type:
        content_type = "application/octet-stream"

    contents = BytesIO(file_data)

    response = make_response(contents.getvalue())
    response.headers.set("Content-Type", content_type)

    is_inline = content_type == "application/pdf"
    disposition = "inline" if is_inline else "attachment"
    response.headers.set("Content-Disposition", disposition, filename=filename)

    return response


def parse_excel_upload(excel_file):
    """Parse an Excel file to extract student data."""
    try:
        df = pd.read_excel(excel_file)
        students = []

        for _, row in df.iterrows():
            student = {
                "first_name": row.get("Name", ""),
                "last_name": row.get("Surname", ""),
                "patronymic": row.get("Patronymic", ""),
                "phone": row.get("Phone", ""),
                "address": "",
                "group_name": row.get("Group", ""),
            }

            # Handle birthday date
            birthday = row.get("Birthday")
            if pd.notna(birthday):
                if hasattr(birthday, "to_pydatetime"):
                    student["birthday_date"] = birthday.to_pydatetime()
                else:
                    try:
                        student["birthday_date"] = datetime.strptime(str(birthday), "%Y-%m-%d")
                    except ValueError:
                        student["birthday_date"] = datetime.now()
            else:
                student["birthday_date"] = datetime.now()

            students.append(student)

        return students
    except Exception as e:
        print(f"Error parsing Excel file: {str(e)}")
        return []
