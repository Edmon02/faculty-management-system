import base64
from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import flash, jsonify, make_response, redirect, render_template, request, session, url_for
from openpyxl import Workbook

from app.services.lecturer_service import LecturerService


class LecturerController:
    @staticmethod
    def list_teachers():
        if not session.get("logged_in"):
            return redirect(url_for("auth.login"))

        group_ids = session.get("group_ids")
        lecturer_service = LecturerService()
        lecturer_data = lecturer_service.get_lecturers_by_groups(group_ids)

        # Handle search filtering
        if request.method == "GET" and any(request.args.values()):
            search_params = {"id": request.args.get("id", default="", type=str), "surname": request.args.get("surname", default="", type=str), "phone": request.args.get("phone", default="", type=str)}

            lecturer_data = lecturer_service.filter_lecturers(lecturer_data, search_params)

        # Handle Excel export
        if request.method == "POST" and request.form.get("action") == "Download":
            return LecturerController._generate_excel(lecturer_data)

        return render_template("teachers.html", data=lecturer_data)

    @staticmethod
    def _generate_excel(lecturer_data):
        # Create Excel workbook and sheet
        wb = Workbook()
        sheet = wb.active

        # Add column headers
        sheet["A1"] = "ID"
        sheet["B1"] = "Name"
        sheet["C1"] = "Surname"
        sheet["D1"] = "Email"
        sheet["E1"] = "Degree"
        sheet["F1"] = "Group Name"
        sheet["G1"] = "Subject Name"

        # Add data to sheet
        for i, data in enumerate(lecturer_data, start=2):
            sheet.cell(row=i, column=1, value=data["id"])
            sheet.cell(row=i, column=2, value=data["first_name"])
            sheet.cell(row=i, column=3, value=data["last_name"])
            sheet.cell(row=i, column=4, value=data["email"])
            sheet.cell(row=i, column=5, value=data["academic_degree"])
            sheet.cell(row=i, column=6, value=data["group_name"])
            sheet.cell(row=i, column=7, value=data["subject_name"])

        # Create response with Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = "attachment; filename=teachers.xlsx"

        return response

    @staticmethod
    def add_teacher():
        if request.method == "POST":
            lecturer_service = LecturerService()

            lecturer_data = {
                "first_name": request.form["first_name"],
                "last_name": request.form["last_name"],
                "patronymic": request.form.get("patronymic", ""),
                "academic_degree": request.form["academic_degree"],
                "position": request.form["position"],
                "email": request.form["email"],
                "phone": request.form.get("phone", ""),
                "is_Admin": request.form.get("is_Admin", "0"),
                "is_Lecturer": request.form.get("is_Lecturer", "1"),
            }

            # Handle image upload
            if "photo" in request.files and request.files["photo"].filename:
                photo = request.files["photo"]
                lecturer_data["images"] = photo.read()

            lecturer_service.add_lecturer(lecturer_data)
            flash("Teacher added successfully", "success")
            return redirect(url_for("teachers.list"))

        return render_template("add-teacher.html")

    @staticmethod
    def edit_teacher(id):
        lecturer_service = LecturerService()

        if request.method == "POST":
            lecturer_data = {
                "last_name": request.form["last_name"],
                "first_name": request.form["first_name"],
                "academic_degree": request.form["academic_degree"],
                "position": request.form["position"],
                "email": request.form["email"],
                "is_Admin": request.form.get("is_Admin", "0"),
                "is_Lecturer": request.form.get("is_Lecturer", "1"),
            }

            # Handle image upload
            if "photo" in request.files and request.files["photo"].filename:
                photo = request.files["photo"]
                lecturer_data["images"] = photo.read()

            lecturer_service.update_lecturer(id, lecturer_data)
            flash("Teacher updated successfully", "success")
            return redirect(url_for("teachers.list"))

        lecturer = lecturer_service.get_lecturer_by_id(id)
        return render_template("edit-teacher.html", data=lecturer)
