# app/routes/teachers.py
from io import BytesIO

from flask import Blueprint, jsonify, make_response, redirect, render_template, request, session, url_for
from openpyxl import Workbook

from app.services.lecturer_service import LecturerService
from app.utils.helpers import login_required

teachers_bp = Blueprint("teachers", __name__)


@teachers_bp.route("/teachers", methods=["GET", "POST"])
@login_required
def teachers_list():
    group_ids = session["group_ids"]
    lecturer_service = LecturerService()

    lecturer_data = lecturer_service.get_all_lecturers_by_groups(group_ids)

    if request.method == "GET":
        search_params = {}

        # Extract query parameters and update search_params dictionary
        search_params["id"] = request.args.get("id", default="", type=str)
        search_params["surname"] = request.args.get("surname", default="", type=str)
        search_params["phone"] = request.args.get("phone", default="", type=str)

        # Filter lecturer_data based on search_params
        if search_params["id"]:
            lecturer_data = [data for data in lecturer_data if search_params["id"] in str(data["id"])]
        if search_params["surname"]:
            lecturer_data = [data for data in lecturer_data if search_params["surname"].lower() == data["last_name"].lower()]
        if search_params["phone"]:
            lecturer_data = [data for data in lecturer_data if search_params["phone"] in data["phone"]]

    if request.method == "POST" and request.form["action"] == "Download":
        # Create a new Excel workbook and sheet
        wb = Workbook()
        sheet = wb.active

        # Add column headers to the sheet
        sheet["A1"] = "ID"
        sheet["B1"] = "Name"
        sheet["C1"] = "Surname"
        sheet["D1"] = "Email"
        sheet["E1"] = "Degree"
        sheet["F1"] = "Group Name"
        sheet["G1"] = "Subject Name"

        # Add data to the sheet
        for i, data in enumerate(lecturer_data, start=2):
            sheet.cell(row=i, column=1, value=data["id"])
            sheet.cell(row=i, column=2, value=data["first_name"])
            sheet.cell(row=i, column=3, value=data["last_name"])
            sheet.cell(row=i, column=4, value=data["email"])
            sheet.cell(row=i, column=5, value=data["academic_degree"])
            sheet.cell(row=i, column=6, value=data["group_name"])
            sheet.cell(row=i, column=7, value=data["subject_name"])

        # Create a bytes buffer to save the workbook to
        buffer = BytesIO()

        # Save the workbook to the buffer
        wb.save(buffer)

        # Set the buffer's seek pointer to the beginning
        buffer.seek(0)

        # Create a response object with the workbook as the content and headers to trigger a file download
        response = make_response(buffer.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = "attachment; filename=teachers.xlsx"

        return response

    return render_template("teachers.html", data=lecturer_data)


@teachers_bp.route("/teachers/add-teacher", methods=["GET", "POST"])
@login_required
def add_teacher():
    return render_template("add-teacher.html")


@teachers_bp.route("/teachers/edit-teacher/<int:id>", methods=["GET", "POST"])
@login_required
def edit_teacher(id):
    lecturer_service = LecturerService()
    data_cursor = lecturer_service.get_lecturer_by_id(id)

    if request.method == "POST":
        search_params = dict(data_cursor)

        # Extract form data and update search_params dictionary
        search_params["last_name"] = request.form["last_name"]
        search_params["first_name"] = request.form["first_name"]
        search_params["academic_degree"] = request.form["academic_degree"]
        search_params["position"] = request.form["position"]
        search_params["email"] = request.form["email"]
        search_params["images"] = request.form["images"]
        search_params["is_Admin"] = request.form.get("is_Admin", default=data_cursor["is_Admin"], type=str)
        search_params["is_Lecturer"] = request.form.get("is_Lecturer", default=data_cursor["is_Lecturer"], type=str)

        # Update the lecturer record in the database
        lecturer_service.update_lecturer(id, search_params, data_cursor)

        # Redirect the user to the teacher list page
        return redirect(url_for("teachers.teachers_list"))

    return render_template("edit-teacher.html", data=data_cursor)
